from openpyxl import workbook, load_workbook
import pandas as pd
import numpy as np

batchExcel_1 = pd.read_excel("data/Batch UPC Data.xlsx")
# batchColumn_2 = set(batchExcel_1.iloc[:, 1])
batchColumn_UPC = batchExcel_1.iloc[:, 1]
batchColumn_Model = batchExcel_1.iloc[:, 2]
batchColumn_Category = batchExcel_1.iloc[:, 4]
batchColumn_Subcategory = batchExcel_1.iloc[:, 5]
batchColumn_Brand = batchExcel_1.iloc[:, 6]

sampleExcel_2 = pd.read_excel("data/Sample UPC Data.xlsx")
# sampleColumn_2 = set(sampleExcel_2.iloc[:, 1])
sampleColumn_UPC = sampleExcel_2.iloc[:, 1]
sampleColumn_Model = sampleExcel_2.iloc[:, 2]
sampleColumn_Category = sampleExcel_2.iloc[:, 4]
sampleColumn_Subcategory = sampleExcel_2.iloc[:, 5]
sampleColumn_Brand = sampleExcel_2.iloc[:, 6]

workbook1 = load_workbook("data/Batch UPC Data.xlsx")
sheet1 = workbook1.active

upc = workbook1["Confirmed_UPC"]

for index, row in batchExcel_1.iterrows():
    value = row["UPC"]
    model = row["Model Number"]
    category = row["Category"]
    subcategory = row["Subcategory"]
    brand = row["Brand"]
    match_found = False
    column_letter = 'N'
    column_letter1 = 'O'
    if value in sampleColumn_UPC.values and model in sampleColumn_Model.values and category in sampleColumn_Category.values and subcategory in sampleColumn_Subcategory.values and brand in sampleColumn_Brand.values:
        match_found = True
        print(f"UPC '{value}' found in both Excel files.")
        row_number = batchExcel_1[batchExcel_1["UPC"]
                                  == value].index.to_numpy()
        row_number = str(row_number)[1:-1]
        row_number = int(row_number) + 2
        cell = sheet1[column_letter + str(row_number)]
        cell.value = "Approved"
        filename = "Verification_Report2.xlsx"
        workbook1.save(filename)
    else:
        print(f"UPC '{value}' invalid")
        row_number = batchExcel_1[batchExcel_1["UPC"]
                                  == value].index.to_numpy()
        row_number = str(row_number)[1:-1]
        row_number = int(row_number) + 2
        cell = sheet1[column_letter + str(row_number)]
        cell.value = "Denied"
        cell = sheet1[column_letter1 + str(row_number)]
        cell.value = "Invalid/Not Found"
        filename = "Verification_Report2.xlsx"
        workbook1.save(filename)

if not match_found:
    print("No values found in the second Excel file.")
# for value in batchColumn_2:
#     column_letter = 'N'
#     column_letter1 = 'O'
#     if value in sampleColumn_2:
#         match_found = True
#         print(f"UPC '{value}' found in both Excel files.")
#         row_number = batchExcel_1[batchExcel_1["UPC"]
#                                   == value].index.to_numpy()
#         row_number = str(row_number)[1:-1]
#         row_number = int(row_number) + 2
#         cell = sheet1[column_letter + str(row_number)]
#         cell.value = "Approved"
#         filename = "Verification_Report2.xlsx"
#         workbook1.save(filename)
#     else:
#         print(f"UPC '{value}' invalid")
#         row_number = batchExcel_1[batchExcel_1["UPC"]
#                                   == value].index.to_numpy()
#         row_number = str(row_number)[1:-1]
#         row_number = int(row_number) + 2
#         cell = sheet1[column_letter + str(row_number)]
#         cell.value = "Denied"
#         cell = sheet1[column_letter1 + str(row_number)]
#         cell.value = "UPC Invalid/Not Found"
#         filename = "Verification_Report2.xlsx"
#         workbook1.save(filename)

# if not match_found:
#     print("No values found in the second Excel file.")

# row_number = batchExcel_1[batchExcel_1.iloc[:, 1]].index.to_numpy()
    # row_number = dataset1[dataset1["UPC"] == batchExcel_1[batchExcel_1.iloc[:, 1]]].index.to_list()
    # row_number = batchExcel_1[batchExcel_1["UPC"]] == sampleExcel_2[sampleExcel_2["UPC"]].index.to_list()
    # row_number = str(row_number)[1:-1]
    # row_number = int(row_number) + 2

# match_found = False
# for value in batchColumn_2:
#     if value in sampleColumn_2:
#         match_found = True
#         print(f"UPC '{value}' found in both Excel files.")
#     else:
#         print(f"UPC '{value}' invalid")

# if not match_found:
#     print("No values found in the second Excel file.")

# for value in batchColumn_2:
#     if value in sampleColumn_2:
#         print(f"UPC '{value}' found")
#     else:
#         print(f"UPC '{value}' not found")

# sampleUPCList = dataset1["UPC"].tolist()
# batchUPCList = dataset2["UPC"].tolist()
# UPC1 = dataset1["UPC"]
# UPC2 = dataset2["UPC"]
# approved_Denied = dataset1['Approve/Denial Status']
# dataset1.set_index("Item ID")

# workbook1 = load_workbook("data/Batch UPC Data.xlsx")
# sheet1 = workbook1.active

# upc = workbook1["Confirmed_UPC"]

# code = int(input("Enter your code:"))

# column_letter = 'P'
# row_number = dataset2[dataset2["UPC"] == code].index.to_list()
# row_number = str(row_number)[1:-1]
# row_number = int(row_number) + 2
# print(row_number)

# x = 622356532419

# for value in batchColumn_2:  # Batch
#     if value in sampleColumn_2:  # Sample
#         print(f"UPC '{value}' found in both Excel files.")
#     else:
#         print(f"UPC '{value}' not found in the second Excel file.")

# for a in batchUPCList:
#     for b in sampleUPCList:
#         if a == b:
#             print("Found")
#         elif a != b:
#             print("Not Found")
# while column_letter == column_letter:
#     if batchUPCList in sampleUPCList == True:
#         print("Found")
    # cell = sheet1[column_letter + str(row_number)]
    # cell.value = "Approved"
    # filename = "Verification_Report2.xlsx"
    # workbook1.save(filename)
    # else:
    #     print("Not Found")
    # cell = sheet1[column_letter + str(row_number)]
    # cell.value = "Denied"
    # filename = "Verification_Report2.xlsx"
    # workbook1.save(filename)

# for u in UPC1:
#     if code == u:
#         for num in UPC2:
#             row_number = dataset2[dataset2["UPC"] == code].index.to_list()
#             row_number = str(row_number)[1:-1]
#             row_number = int(row_number) + 2
#             if code == num:
#                 print("Success")
#                 cell = sheet1[column_letter + str(row_number)]
#                 cell.value = "Approved"
#                 filename = "Verification_Report2.xlsx"
#                 workbook1.save(filename)
#             else:
#                 print("Failure")
#                 cell = sheet1[column_letter + str(row_number)]
#                 cell.value = "Denied"
#                 filename = "Verification_Report2.xlsx"
#                 workbook1.save(filename)
    # first_element = [u]
    # print("Success")
    # id = dataset1[dataset1['Approve/Denial Status'] == code].index
    # dataset1.loc[id:"Approve/Denial Status"] == "Approved"
    # print("Approved")
    # dataset1["Approve/Denial Status"] == "Approved"
    # else:
    #     print("Next")
