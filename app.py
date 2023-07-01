import os

from flask import Flask, render_template, request, send_file, make_response

from fileinput import filename

import pandas as pd

from openpyxl import workbook, load_workbook

app = Flask(__name__)

# Individual UPC Verification


@app.route('/', methods=["GET", "POST"])
def form():
    if request.method == "POST" and "z" in request.form:
        x = request.form.get('z')
        x = int(x)
        dataset = pd.read_excel("data/Sample UPC Data.xlsx")
        upc1 = dataset["UPC"]
        workbook1 = load_workbook("data/Sample UPC Data.xlsx")
        sheet1 = workbook1.active
        upc = workbook1["Confirmed_UPC"]
        column_letter = 'P'
        row_number = dataset[dataset["UPC"] == x].index.to_list()
        row_number = str(row_number)[1:-1]
        row_number = int(row_number) + 2
        filename = "Verification_Report.xlsx"
        for u in upc1:
            if x == u:
                cell = sheet1[column_letter + str(row_number)]
                cell.value = "Approved"
                filename = "Verification_Report.xlsx"
                workbook1.save(filename)
        return "Data has been Updated Successfully in Verification_Report.excel!! Check it out!!"
    return render_template('verification.html')


@app.route('/uploadVerification')
def uploadVerification():
    return render_template('verification.html')


@app.route('/download')
def download():
    file_path = r"C:\Users\kidus\myflaskproject\data\UPCUploadTemplate.xlsx"
    filename = "UPCUploadTemplate.xlsx"
    # Create the response object
    response = make_response(send_file(file_path))
    # Set the "Content-Disposition" header to trigger file download
    response.headers["Content-Disposition"] = "attachment; filename=" + filename
    # Get the user's home directory
    home_dir = os.path.expanduser("~")
    # Set the path to the "Downloads" folder
    downloads_folder = os.path.join(home_dir, "Downloads")
    # Set the "X-Sendfile" header to suggest the default save location
    response.headers["X-Sendfile"] = os.path.join(downloads_folder, filename)
    return response


@app.route('/success', methods=['POST'])
def success():
    if request.method == 'POST':
        # Handling uploaded file
        f = request.files['file']
        f.save(f.filename)
        uploadedFile = f
        # Reading Necessary Excel Files
        batchExcel_1 = pd.read_excel(uploadedFile)
        batchExcel_1 = pd.read_excel("data/Batch UPC Data.xlsx")
        # batchColumn_2 = set(batchExcel_1.iloc[:, 1])
        # Initializing Key Variables from Imported Excel File
        batchColumn_UPC = batchExcel_1.iloc[:, 1]
        batchColumn_Model = batchExcel_1.iloc[:, 2]
        batchColumn_Category = batchExcel_1.iloc[:, 4]
        batchColumn_Subcategory = batchExcel_1.iloc[:, 5]
        batchColumn_Brand = batchExcel_1.iloc[:, 6]
        # Reading Excel File Saved on Local Machine
        sampleExcel_2 = pd.read_excel("data/Sample UPC Data.xlsx")
        # sampleColumn_2 = set(sampleExcel_2.iloc[:, 1])
        # Initializing Key Variables from Imported Excel File
        sampleColumn_UPC = sampleExcel_2.iloc[:, 1]
        sampleColumn_Model = sampleExcel_2.iloc[:, 2]
        sampleColumn_Category = sampleExcel_2.iloc[:, 4]
        sampleColumn_Subcategory = sampleExcel_2.iloc[:, 5]
        sampleColumn_Brand = sampleExcel_2.iloc[:, 6]

        # Loading the Workbook and Setting the Active Sheet
        workbook1 = load_workbook("data/Batch UPC Data.xlsx")
        sheet1 = workbook1.active

        upc = workbook1["Confirmed_UPC"]

        # Identifies the Row Number and the Iterates through the Content in the Uploaded File
        for index, row in batchExcel_1.iterrows():
            value = row["UPC"]
            model = row["Model Number"]
            category = row["Category"]
            subcategory = row["Subcategory"]
            brand = row["Brand"]
            match_found = False
            column_letter = 'N'
            column_letter1 = 'O'
            # Checks to See if the Values in the Excel File Match the Ones in the Sample Data Excel File
            if value in sampleColumn_UPC.values and model in sampleColumn_Model.values and category in sampleColumn_Category.values and subcategory in sampleColumn_Subcategory.values and brand in sampleColumn_Brand.values:
                # If there is a Match, the Row Number of the Item will be Found and "Approved" will be written in the "Approve/Denial" Column
                match_found = True
                print(f"UPC '{value}' found in both Excel files.")
                row_number = batchExcel_1[batchExcel_1["UPC"]
                                          == value].index.to_numpy()
                row_number = str(row_number)[1:-1]
                row_number = int(row_number) + 2
                cell = sheet1[column_letter + str(row_number)]
                cell.value = "Approved"
                # Outputs Result to an Excel File that is Downloaded
                filename = "UPC Verification Report.xlsx"
                workbook1.save(filename)
            else:
                # If There is no Match, the Row Number of the Item will be Found and "Denied" will be written in the "Approve/Denial" Column as well as the Denial Reason
                print(f"UPC '{value}' invalid")
                row_number = batchExcel_1[batchExcel_1["UPC"]
                                          == value].index.to_numpy()
                row_number = str(row_number)[1:-1]
                row_number = int(row_number) + 2
                cell = sheet1[column_letter + str(row_number)]
                cell.value = "Denied"
                cell = sheet1[column_letter1 + str(row_number)]
                cell.value = "Invalid/Not Found"
                # Outputs Result to an Excel File that is Downloaded
                filename = "UPC Verification Report.xlsx"
                workbook1.save(filename)

        if not match_found:
            print("No values found in the second Excel file.")
        # batchColumn_UPC = batchExcel_1.iloc[:, 1]
        # batchColumn_Model = batchExcel_1.iloc[:, 2]
        # batchColumn_Category = batchExcel_1.iloc[:, 4]
        # batchColumn_Subcategory = batchExcel_1.iloc[:, 5]
        # batchColumn_Brand = batchExcel_1.iloc[:, 6]

        # sampleExcel_2 = pd.read_excel("data/Sample UPC Data.xlsx")
        # sampleColumn_UPC = sampleExcel_2.iloc[:, 1]
        # sampleColumn_Model = sampleExcel_2.iloc[:, 2]
        # sampleColumn_Category = sampleExcel_2.iloc[:, 4]
        # sampleColumn_Subcategory = sampleExcel_2.iloc[:, 5]
        # sampleColumn_Brand = sampleExcel_2.iloc[:, 6]

        # workbook1 = load_workbook(uploadedFile)
        # sheet1 = workbook1.active

        # upc = workbook1["Confirmed_UPC"]

        # match_found = False
        # for index, value in enumerate(batchColumn_UPC):
        #     approve_deny_column_letter = 'N'
        #     reason_column_letter = 'O'
        #     model = batchColumn_Model[index]
        #     category = batchColumn_Category[index]
        #     subcategory = batchColumn_Subcategory[index]
        #     brand = batchColumn_Brand[index]
        #     if value in sampleColumn_UPC and model in sampleColumn_Model and category in sampleColumn_Category and subcategory in sampleColumn_Subcategory and brand in sampleColumn_Brand:
        #         match_found = True
        #         # print(f"UPC '{value}' found in both Excel files.")
        #         row_number = batchExcel_1[batchExcel_1["UPC"]
        #                                   == value].index.to_numpy()
        #         row_number = str(row_number)[1:-1]
        #         row_number = int(row_number) + 2
        #         cell = sheet1[approve_deny_column_letter + str(row_number)]
        #         cell.value = "Approved"
        #         outputFileName = "Verification_Report2.xlsx"
        #         workbook1.save(outputFileName)
        #     else:
        #         # print(f"UPC '{value}' invalid")
        #         row_number = batchExcel_1[batchExcel_1["UPC"]
        #                                   == value].index.to_numpy()
        #         row_number = str(row_number)[1:-1]
        #         row_number = int(row_number) + 2
        #         cell = sheet1[approve_deny_column_letter + str(row_number)]
        #         cell.value = "Denied"
        #         cell = sheet1[reason_column_letter + str(row_number)]
        #         cell.value = "UPC Invalid/Not Found"
        #         outputFileName = "Verification_Report2.xlsx"
        #         workbook1.save(outputFileName)

        # if not match_found:
        #     print("No values found in the second Excel file.")

        return render_template('checkout.html')
    return render_template('verification.html')


if __name__ == "__main__":
    app.run(debug=True)

#  host='0.0.0.0', port=5000, server_name='GS1 Scout'
